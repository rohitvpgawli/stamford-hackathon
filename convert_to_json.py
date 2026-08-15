"""
Converts the filled Mango_Stamford_Dataset_Template.xlsx into opportunities.json
matching the exact schema in Rohit's PRD (section 8.5).

Usage:
    python3 convert_to_json.py Mango_Stamford_Dataset_Template.xlsx opportunities.json
"""
import sys
import json
import uuid
import re
from datetime import datetime
import openpyxl

CATEGORY_TABS = ["CT", "PR", "UC", "VO", "EV", "LP"]

def slugify(text):
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text.strip().lower()).strip("-")
    return text[:40]

def parse_bool_status(status_raw):
    valid = {"active", "canceled", "sold_out", "expired"}
    s = (status_raw or "active").strip().lower()
    return s if s in valid else "active"

def combine_datetime(date_str, time_str):
    if not date_str or "TBD" in str(date_str).upper():
        return None
    date_str = str(date_str).strip()
    time_str = str(time_str).strip() if time_str else "00:00"
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            if " " not in fmt or time_str:
                combined = f"{date_str} {time_str}" if time_str else date_str
                return datetime.strptime(combined.strip(), "%Y-%m-%d %H:%M").isoformat()
        except ValueError:
            continue
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").isoformat()
    except ValueError:
        return None

def row_to_dict(headers, row_values):
    return dict(zip(headers, row_values))

def convert(xlsx_path, json_path):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    opportunities = []
    skipped = 0

    for tab in CATEGORY_TABS:
        if tab not in wb.sheetnames:
            print(f"WARNING: tab {tab} not found, skipping")
            continue
        ws = wb[tab]

        header_row_idx = 6
        headers = []
        for col in range(2, 28):
            val = ws.cell(row=header_row_idx, column=col).value
            headers.append(val)

        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(2, 28)]
            if not any(row_values):
                continue
            rec = row_to_dict(headers, row_values)

            title = (rec.get("title") or "").strip()
            if not title or "EXAMPLE" in str(rec.get("researcher_notes", "")).upper():
                continue
            if title.startswith("TODO"):
                skipped += 1
                continue

            tags = [t.strip() for t in str(rec.get("tags") or "").split(",") if t.strip()]
            audience = [a.strip() for a in str(rec.get("audience") or "").split(",") if a.strip()]

            price_usd = rec.get("price_usd")
            try:
                price_cents = int(float(price_usd) * 100) if price_usd not in (None, "", "TODO") else 0
            except (ValueError, TypeError):
                price_cents = 0

            starts_at = combine_datetime(rec.get("event_date"), rec.get("start_time"))
            ends_at = combine_datetime(rec.get("event_date"), rec.get("end_time")) if rec.get("end_time") else None

            opportunity = {
                "id": str(uuid.uuid4()),
                "kind": (rec.get("kind") or "event").strip().lower(),
                "title": title,
                "short_description": (rec.get("short_description") or "").strip(),
                "venue_name": (rec.get("venue_name") or "").strip(),
                "neighborhood": (rec.get("neighborhood") or "").strip(),
                "latitude": rec.get("latitude") or None,
                "longitude": rec.get("longitude") or None,
                "starts_at": starts_at,
                "ends_at": ends_at,
                "recurring_text": (rec.get("recurring_text") or "").strip() or None,
                "price_cents": price_cents,
                "capacity": rec.get("capacity") or None,
                "status": parse_bool_status(rec.get("status")),
                "source_name": (rec.get("source_name") or "").strip(),
                "source_url": (rec.get("source_url") or "").strip(),
                "is_demo_data": True,
                "tags": tags,
                "audience": audience,
                "group_style": (rec.get("group_style") or "").strip() or None,
                "accessibility": {"notes": rec.get("accessibility")} if rec.get("accessibility") else None,
                "transport_notes": (rec.get("transport_notes") or "").strip() or None,
                "photo_url": (rec.get("photo_url") or "").strip() or None,
                "photo_source": (rec.get("photo_source") or "").strip() or None,
                "category_source": tab,
            }
            opportunities.append(opportunity)

    with open(json_path, "w") as f:
        json.dump(opportunities, f, indent=2, default=str)

    print(f"Wrote {len(opportunities)} opportunities to {json_path}")
    if skipped:
        print(f"Skipped {skipped} rows still marked TODO (incomplete) — finish these before demo day")

    by_cat = {}
    for o in opportunities:
        by_cat[o["category_source"]] = by_cat.get(o["category_source"], 0) + 1
    print("Breakdown by category:", by_cat)

if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "Mango_Stamford_Dataset_Template.xlsx"
    json_path = sys.argv[2] if len(sys.argv) > 2 else "opportunities.json"
    convert(xlsx_path, json_path)
